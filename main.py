#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Steam maFile scanner → bans fetcher → JSON + XLSX report
З обмеженням частоти запитів 10–15 на хвилину (за замовчуванням).

ЩО РОБИТЬ СКРИПТ
1) Зчитує всі .maFile у вказаній теці (за замовчуванням ./maFile),
   дістає SteamID та логін (назва файла без розширення).
2) Пакетами звертається до Steam Web API GetPlayerBans.
3) Зберігає:
   - steam_ids64.txt          — список SteamID
   - data.json                — «сирі» відповіді API (players)
   - data_ban.json            — відфільтровані з банами/свіжими банами
   - report.xlsx              — Excel: Login, SteamID, Ігрові блокування, Ком'юніті бан, VAC бан
4) Має простий CLI: --dir, --chunk, --days, --out-xlsx, та контроль RPM: --rpm-min/--rpm-max.

НЕОБХІДНО:
- Python 3.8+
- pip install requests openpyxl
- Ключ API в змінній середовища STEAM_API_KEY або у файлі .env рядком:
    STEAM_API_KEY=ваш_ключ

ПРИКЛАД ЗАПУСКУ:
    python main.py --dir maFile --chunk 50 --days 10 --out-xlsx report.xlsx
    # зміна ліміту:
    python main.py --rpm-min 10 --rpm-max 15
"""

from __future__ import annotations

import argparse
import json
import os
import random
import time
from pathlib import Path
from typing import Dict, Iterable, Iterator, List, Tuple

import requests
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ------------------------- Обмежувач запитів ------------------------- #

class MinuteRateLimiter:
    """
    Підтримує середню частоту запитів у діапазоні [min_rpm, max_rpm] на хвилину.
    Після кожного виклику wait() встановлюється випадкова пауза між:
        [60/max_rpm, 60/min_rpm] секунд.
    Таким чином запити рівномірно «розмазуються» у межах 10–15/хв (за замовч.).
    """
    def __init__(self, min_rpm: int = 10, max_rpm: int = 15):
        if min_rpm <= 0 or max_rpm <= 0 or max_rpm < min_rpm:
            raise ValueError("Некоректні значення RPM. Очікується 0 < min_rpm ≤ max_rpm.")
        self.min_interval = 60.0 / float(max_rpm)  # найменша пауза (для верхньої межі)
        self.max_interval = 60.0 / float(min_rpm)  # найбільша пауза (для нижньої межі)
        self._next_earliest = 0.0

    def wait(self) -> None:
        now = time.monotonic()
        if now < self._next_earliest:
            time.sleep(self._next_earliest - now)
        interval = random.uniform(self.min_interval, self.max_interval)
        self._next_earliest = time.monotonic() + interval


# ------------------------- Утиліти ------------------------- #

def load_env_dotenv(dotenv_path: Path = Path(".env")) -> None:
    """Проста (без залежностей) підтримка .env: завантажує STEAM_API_KEY та інші, якщо знайде."""
    if not dotenv_path.exists():
        return
    for line in dotenv_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if not line or line.startswith("#") or "=" not in line:
            continue
        k, v = line.split("=", 1)
        k = k.strip()
        v = v.strip().strip('"').strip("'")
        if k and v and (k not in os.environ):
            os.environ[k] = v


def chunked(it: Iterable[str], size: int) -> Iterator[List[str]]:
    buf: List[str] = []
    for x in it:
        buf.append(x)
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


def read_mafiles(ma_dir: Path) -> List[Tuple[str, str]]:
    """
    Читає всі *.maFile, повертає список пар (login, steamid).
    login — це ім'я файла без розширення.
    """
    pairs: List[Tuple[str, str]] = []
    for fp in sorted(ma_dir.glob("*.maFile")):
        try:
            data = json.loads(fp.read_text(encoding="utf-8"))
        except Exception:
            data = json.loads(fp.read_text(encoding="utf-8", errors="ignore"))

        # Найчастіше SteamID знаходиться у Session.SteamID
        steamid = None
        if isinstance(data, dict):
            if "Session" in data and isinstance(data["Session"], dict):
                steamid = data["Session"].get("SteamID") or data["Session"].get("steamid")
            steamid = steamid or data.get("SteamID") or data.get("steamid") or data.get("steam_id")

        if not steamid:
            print(f"⚠️  Не знайшов SteamID у файлі: {fp.name}")
            continue

        steamid_str = str(steamid).strip()
        if not steamid_str.isdigit():
            print(f"⚠️  Підозрілий SteamID у {fp.name}: {steamid_str}")
            continue

        login = fp.stem
        pairs.append((login, steamid_str))

    return pairs


def fetch_player_bans(
    api_key: str,
    steam_ids: List[str],
    *,
    rate_limiter: MinuteRateLimiter | None = None,
    timeout: int = 20,
    retries: int = 3,
    backoff_sec: float = 1.5,
) -> List[dict]:
    """
    Витягує бан-метадані для списку SteamID через ISteamUser/GetPlayerBans.
    Повертає список словників (players).
    ВАЖЛИВО: перед кожним HTTP-запитом викликається rate_limiter.wait() (якщо передано).
    """
    url = "https://api.steampowered.com/ISteamUser/GetPlayerBans/v1/"
    params = {"key": api_key, "steamids": ",".join(steam_ids)}

    for attempt in range(1, retries + 1):
        try:
            if rate_limiter:
                rate_limiter.wait()  # обмеження 10–15 запитів/хв
            resp = requests.get(url, params=params, timeout=timeout)
            if resp.status_code == 429:
                # Ліміт — почекаємо й повторимо (не забуваємо: наступний запит теж пройде через wait()).
                wait = backoff_sec * attempt
                print(f"⏳ Rate limit (429). Чекаю {wait:.1f}s і повторюю...")
                time.sleep(wait)
                continue
            resp.raise_for_status()
            payload = resp.json()
            players = payload.get("players") or payload.get("Players") or []
            if not isinstance(players, list):
                players = []
            return players
        except requests.RequestException as e:
            if attempt >= retries:
                print(f"❌ HTTP помилка: {e}")
                raise
            wait = backoff_sec * attempt
            print(f"⚠️  Помилка мережі: {e}. Повтор через {wait:.1f}s...")
            time.sleep(wait)

    return []


def save_json(obj: object, path: Path) -> None:
    path.write_text(json.dumps(obj, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"💾 Збережено JSON: {path}")


def save_text_lines(lines: Iterable[str], path: Path) -> None:
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    print(f"💾 Збережено список: {path}")


def bool_to_uk(v: bool) -> str:
    return "Так" if bool(v) else "Ні"


def save_to_xlsx(players: List[dict], steamid_to_login: Dict[str, str], out_path: Path) -> None:
    """
    Створює XLSX зі стовпцями:
    Login | SteamID | Ігрові блокування | Ком'юніті бан | VAC бан
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Bans"

    headers = ["Login", "SteamID", "Ігрові блокування", "Ком'юніті бан", "VAC бан"]
    ws.append(headers)

    for p in players:
        sid = str(p.get("SteamId", "") or p.get("steamid", ""))
        login = steamid_to_login.get(sid, "Невідомо")
        row = [
            login,
            sid,
            int(p.get("NumberOfGameBans", 0) or 0),
            bool_to_uk(p.get("CommunityBanned", False)),
            bool_to_uk(p.get("VACBanned", False)),
        ]
        ws.append(row)

    # Автоширина
    for col_idx in range(1, len(headers) + 1):
        col = get_column_letter(col_idx)
        max_len = max(
            len(str(ws[f"{col}{r}"].value)) if ws[f"{col}{r}"].value is not None else 0
            for r in range(1, ws.max_row + 1)
        )
    # Трішки ширше для зручності читання
        ws.column_dimensions[col].width = min(max(12, max_len + 2), 50)

    wb.save(out_path.as_posix())
    print(f"📊 Збережено Excel: {out_path}")


def make_ban_rows(
    players: List[dict],
    steamid_to_login: Dict[str, str],
    *,
    days_threshold: int | None,
) -> List[dict]:
    """
    Готує відфільтрований список з розширеними полями:
    login, SteamId, profile_url, VACBanned, NumberOfGameBans, CommunityBanned, DaysSinceLastBan
    Фільтрація:
      - якщо days_threshold задано: беруться ті, у кого 0 < DaysSinceLastBan ≤ threshold
      - інакше — усі, в кого VACBanned або NumberOfGameBans>0 або CommunityBanned=True
    """
    out: List[dict] = []
    for p in players:
        sid = str(p.get("SteamId", "") or p.get("steamid", ""))
        dslb = p.get("DaysSinceLastBan")
        vac = bool(p.get("VACBanned", False))
        game_bans = int(p.get("NumberOfGameBans", 0) or 0)
        comm = bool(p.get("CommunityBanned", False))

        take = False
        if days_threshold is not None:
            try:
                d = int(dslb)
                take = (0 < d <= days_threshold)
            except (TypeError, ValueError):
                take = False
        else:
            take = vac or game_bans > 0 or comm

        if take:
            out.append({
                "login": steamid_to_login.get(sid, "Невідомо"),
                "SteamId": sid,
                "profile_url": f"https://steamcommunity.com/profiles/{sid}" if sid else "",
                "VACBanned": vac,
                "NumberOfGameBans": game_bans,
                "CommunityBanned": comm,
                "DaysSinceLastBan": dslb,
            })
    return out


# ------------------------- Головна логіка ------------------------- #

def env_int(name: str, default: int | None) -> int | None:
    v = os.getenv(name)
    if v is None or v == "":
        return default
    try:
        return int(v)
    except ValueError:
        return default


def main() -> None:
    load_env_dotenv()

    parser = argparse.ArgumentParser(description="Scan .maFiles → Steam bans → JSON/XLSX (з лімітом 10–15 запитів/хв)")
    parser.add_argument("--dir", default="maFiles", help="Тека з *.maFile (за замовчуванням: maFile)")
    parser.add_argument("--chunk", type=int, default=50, help="Розмір пакета для API (default: 50)")
    parser.add_argument("--days", type=int, default=None,
                        help="Поріг днів для свіжих банів (напр., 10). Якщо не задано — беруться всі з банами.")
    parser.add_argument("--out-steamids", default="steam_ids64.txt", help="Файл зі SteamID (default: steam_ids64.txt)")
    parser.add_argument("--out-json", default="data.json", help="Файл з повними відповідями API (default: data.json)")
    parser.add_argument("--out-ban", default="data_ban.json",
                        help="Файл з відфільтрованими банами (default: data_ban.json)")
    parser.add_argument("--out-xlsx", default="report.xlsx", help="Excel звіт (default: report.xlsx)")
    parser.add_argument("--rpm-min", type=int, default=env_int("RPM_MIN", 10),
                        help="Мінімальна кількість запитів на хвилину (default/env RPM_MIN: 10)")
    parser.add_argument("--rpm-max", type=int, default=env_int("RPM_MAX", 15),
                        help="Максимальна кількість запитів на хвилину (default/env RPM_MAX: 15)")
    args = parser.parse_args()

    api_key = os.getenv("STEAM_API_KEY", "").strip()
    if not api_key:
        raise SystemExit("❌ Немає ключа API. Задайте STEAM_API_KEY у середовищі або у .env")

    ma_dir = Path(args.dir)
    if not ma_dir.exists():
        raise SystemExit(f"❌ Тека не існує: {ma_dir.resolve()}")

    # 1) зчитуємо maFile
    pairs = read_mafiles(ma_dir)
    if not pairs:
        raise SystemExit("❌ Не знайдено жодного валідного .maFile")

    # дедуплікація за SteamID: перший login перемагає
    steamid_to_login: Dict[str, str] = {}
    for login, sid in pairs:
        if sid not in steamid_to_login:
            steamid_to_login[sid] = login

    steam_ids = list(steamid_to_login.keys())
    print(f"✅ Знайдено SteamID: {len(steam_ids)}")

    # 2) зберігаємо список SteamID
    save_text_lines(steam_ids, Path(args.out_steamids))

    # 3) ініціалізуємо обмежувач
    rate_limiter = MinuteRateLimiter(min_rpm=max(1, args.rpm_min), max_rpm=max(args.rpm_min, args.rpm_max))

    # 4) тягнемо бани пакетами, обмеження — ПЕРЕД КОЖНИМ HTTP-ЗАПИТОМ усередині fetch_player_bans
    all_players: List[dict] = []
    for batch in chunked(steam_ids, max(1, int(args.chunk))):
        players = fetch_player_bans(api_key, batch, rate_limiter=rate_limiter)
        all_players.extend(players)
        print(f"… отримано {len(players)} запис(ів) у батчі, всього: {len(all_players)}")

    # 5) повні відповіді
    save_json(all_players, Path(args.out_json))

    # 6) фільтр для data_ban.json
    ban_rows = make_ban_rows(all_players, steamid_to_login, days_threshold=args.days)
    save_json(ban_rows, Path(args.out_ban))
    print(f"🧮 Відфільтровано записів у data_ban.json: {len(ban_rows)}")

    # 7) Excel (колонки: Login, SteamID, Ігрові блокування, Ком'юніті бан, VAC бан)
    save_to_xlsx(all_players, steamid_to_login, Path(args.out_xlsx))

    print("🎉 Готово!")


if __name__ == "__main__":
    main()
